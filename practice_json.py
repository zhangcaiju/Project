import yaml
from openpyxl import Workbook
import os

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("Mysheet")
ws2 = wb.create_sheet("Mysheet", 0)
ws3 = wb.create_sheet("Mysheet", 1)

ws.title = "New Title"
ws1.title = "1"
ws2.title = "2"
d = ws.cell(row=4, column=2, value=10)
Workbook.active.cell

wb.save("Mysheet.xlsx")
# wb.remove_sheet(wb.get_sheet_by_name("Mysheet"))

